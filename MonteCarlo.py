import pandas as pd
import numpy as np
from matplotlib import pyplot as plt
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook


weight=pd.read_excel(r"weight.xlsx",header=None)[0].tolist()

# *********************************************************回测部分
Rs=0.06
Rb=0.03
sig_s=0.22
sig_b=0
stockprice={}
bondprice={}
principal={}
for i in range(0,1000):
    S=np.zeros(20)
    B=np.zeros(20)
    P=np.zeros(20)
    S[0]=100
    B[0]=100
    P[0]=1
    for j in range(0,20):
        dS=Rs*S[j]*1+sig_s*S[j]*np.random.randn()*np.sqrt(1)
        dB=Rb*B[j]*1+sig_b*B[j]*np.random.randn()*np.sqrt(1)
        S[j+1]=S[j]+dS
        B[j+1]=B[j]+dB
        P[j+1]=S[j+1]*weight[j+20]*P[j]/S[j]+B[j+1]*(1-weight[j+20])*P[j]/B[j]   
    stockprice[i]=S
    bondprice[i]=B
    principal[i]=P

BondPrice=pd.DataFrame.from_dict(bondprice,orient="index")
StockPrice=pd.DataFrame.from_dict(stockprice,orient="index")
Principal=pd.DataFrame.from_dict(principal,orient="index")
Result_principal=Principal.T
Result_principal.plot(legend=None)
ret=(Result_principal.iloc[-1,:]-Result_principal.iloc[0,:])/Result_principal.iloc[0,:]
ret_summary=ret.describe()
annua_ret=(1+ret)**(1/20)-1
annua_summary=annua_ret.describe()
annua_summary['result']= '平均年化收益率'
annua_summary = annua_summary.to_dict()


lag_principal=Result_principal.shift(1)
ret_annual=(Result_principal-lag_principal)/lag_principal
std_annual=ret_annual.std()
std_summarize=std_annual.describe()
std_summarize['result']=  '平均波动率'
std_summarize = std_summarize.to_dict()

result=pd.DataFrame([[annua_summary['mean'],std_summarize['mean']],[annua_summary['max'],std_summarize['max']],[annua_summary['min'],std_summarize['min']]])
result.index=['mean','max','min']
result.columns=['年化收益率','年化波动率']
wb2 = load_workbook(r"return&std.xlsx")
ws2=wb2.create_sheet("MonteCarlo")
for r in dataframe_to_rows(result, index=True, header=True):
    ws2.append(r)
wb2.save("return&std.xlsx")

plt.hist(annua_ret)